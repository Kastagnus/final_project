import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

def export_reservations_to_excel(queryset):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Reservations'

    # Define the column titles
    columns = ['Reservation Number', 'User', 'Tour Title', 'Start Date', 'End Date', 'Number of People', 'Status', 'Total Price']

    # Write the column titles to the worksheet
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data to the worksheet
    for row_num, reservation in enumerate(queryset, 2):
        worksheet.cell(row=row_num, column=1, value=reservation.reservation_reference)
        worksheet.cell(row=row_num, column=2, value=reservation.user.username)
        worksheet.cell(row=row_num, column=3, value=reservation.tour.title)
        worksheet.cell(row=row_num, column=4, value=reservation.start_date)
        worksheet.cell(row=row_num, column=5, value=reservation.end_date)
        worksheet.cell(row=row_num, column=6, value=reservation.number_of_people)
        worksheet.cell(row=row_num, column=7, value=reservation.status)
        worksheet.cell(row=row_num, column=8, value=reservation.total_price)

    # Adjust column widths
    for col_num, column_title in enumerate(columns, 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reservations.xlsx'
    workbook.save(response)

    return response
